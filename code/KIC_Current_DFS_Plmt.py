import pandas as pd, pyautogui as p, json, numpy as np, os, datetime as dt, re, sys, stuff
from simple_salesforce import Salesforce, format_soql
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def kic_curr_dfs_plmt(df = pd.DataFrame()):

    def column_size(sheet):
        '''Dynamically adjust the column sizes in excel sheet'''
        column_widths = []
        for row in sheet:
            for i, cell in enumerate(row):
                if len(column_widths) > i:
                    if len(str(cell.value)) > column_widths[i]:
                        column_widths[i] = len(str(cell.value))+5
                else:
                    column_widths += [len(str(cell.value))+5]
        for i, column_width in enumerate(column_widths):
            sheet.column_dimensions[get_column_letter(i+1)].width = column_width		

    def soql_df(soql_query):
        '''Create a DF using SOQL query and normalizing JSON'''
        info = sf.query_all(soql_query)
        df = pd.json_normalize(info['records'])
        cols = [c for c in df.columns if 'attribute' not in c]
        cols = [c for c in cols if not c.endswith('__r')]
        df = df[cols].copy()
        return df

    def cus_rank(x):
        '''Rank anything that is not DSCYF/DFS, DFS, DSCYF as 1 for sorting'''
        if x in ('DSCYF/DFS', 'DFS', 'DSCYF'):
            return 2
        else:
            return 1

    def rank_null(x):
        '''Rank null dates above a normal date for sorting'''
        if x == None:
            return 1
        else:
            return 2

    def check_prev(x, prevdf):
        
        if x in list(prevdf['PID']):
            return '--'
        else:
            return 'NEW'

    def color_row(row):
        if row['In Previous Report?'] == 'NEW':
            color = '#fce4d6'
            return [f'background-color: {color}'] * len(row.values)
        
    if df.empty:   #If there is no df then run the report from the beginning.
        #Sign into Salesforce.
        username = stuff.username
        #password = p.password('Enter your password', title='Salesforce Password')
        password = stuff.password
        orgid = ''
        url = ''
        sf = Salesforce(username = username, password = password, instance_url = url, organizationId = orgid)

        #Create the initial Kids in Custody query.
        q = f'''
        SELECT Custodian_Name__c,Start_Date__c,End_Date__c,Client_Name__r.DEL_PID__c,Client_Name__r.Name,Client_Name__r.Birthdate,
        Client_Name__r.DEL_Age__c, Client_Name__r.DEL_Age_Years__c,Client_Name__r.DEL_Custody_Start_Date__c,
        Client_Name__r.DEL_Runaway_Alert_Flag__c, Client_Name__r.DEL_MCI__c
        FROM DEL_Custody__c

        ORDER BY Client_Name__r.DEL_PID__c, Start_Date__c DESC
        '''

        df = soql_df(q)

        #Rank the custodian and null dates
        df['Custodian_Rank'] = df['Custodian_Name__c'].apply(cus_rank)
        df['Date_Rank'] = df['End_Date__c'].apply(rank_null)

        #Sort the data to get the most recent start date with a non DFS custodian first and open end date first as well.
        df = df.sort_values(by=['Client_Name__r.DEL_PID__c', 'Start_Date__c',
                                  'Custodian_Rank', 'Date_Rank'], ascending = [True, False,True, True]).copy()

        #Drop duplicate by PID and then keep the ones that are DFS related with an open end date.
        df = df.drop_duplicates('Client_Name__r.DEL_PID__c').copy()

        df = df[df['Custodian_Name__c'].isin(['DSCYF/DFS', 'DFS', 'DSCYF'])].copy()

        df = df[df['End_Date__c'].isnull()].copy()

        #Drop the rank columns and date columns
        df = df.drop(['Start_Date__c','End_Date__c','Custodian_Rank', 'Date_Rank'], axis = 'columns').copy()

        #Change Age_Years into a numeric.
        df = df.astype({'Client_Name__r.DEL_Age_Years__c' : 'int32'}).copy()

        #Make a list of all the PIDs
        pids = list(df['Client_Name__r.DEL_PID__c'])

        #Create the SOQL query for all the Placements for those PIDs.
        q2=format_soql('''
        SELECT PID__c, Name,Division__c, Service_Name__c,
            Case_Number__r.Assigned_Worker__r.Name , Case_Number__r.Assigned_Worker__r.DEL_Service_Area__c,
            Case_Number__r.Assigned_Supervisor__r.Name, State__c, Placement_Start_Date_Division_Wide__c,
            Case_Number__r.Name, Case_Type__c, Placement_Type_Formula__c
        FROM DEL_Placement__c
        WHERE PID__c IN {pids}
        AND Service_Name__c != null
        AND Placement_Start_Date_Division_Wide__c != null
        AND Placement_End_Date_Division_Wide__c = null
        AND Latest_Version_Placement__c = True
        ''', pids = pids)

        #Run the query and put it into the Pandas DataFrame.
        df2 = soql_df(q2)

        #Merge df1 with df2 and rename columns and order them correctly.
        df3 = df.merge(df2, how='left',left_on = 'Client_Name__r.DEL_PID__c', right_on = 'PID__c').copy()
        df3 = df3.drop(columns = ['PID__c']).copy()

        columns = [
            'Who Has Custody','PID', 'Name', 'Birth Date', 'Age', 'Age (Years)', 'Custody Start Date', 'Runaway',
            'MCI','Placement ID', 'Division', 'Service', 'State', 'Placement Start Date', 'Case Type','Placement Type' ,'Case Owner',
            'Service Area', 'Supervisor', 'Case Number'
        ]

        df3.columns = columns #Rename columns

        ord_columns = [
            'PID', 'Name', 'Birth Date', 'Age (Years)', 'Custody Start Date','Placement Start Date',
            'Case Type','Case Owner','Supervisor','Service Area','Division','Case Number',  
            'Placement ID','Placement Type'
        ]

        df3 = df3[ord_columns].copy() #Order the columns

        #Kids in Custody with Active DFS Placement
        wplacement = df3.loc[~df3['Placement ID'].isnull()].copy()
        wplacement = wplacement[wplacement['Division']=='DFS'].copy()
        wplacement.loc[:,['Birth Date','Custody Start Date', 'Placement Start Date']]= wplacement.loc[:,['Birth Date','Custody Start Date', 'Placement Start Date']].apply(pd.to_datetime) 
        wplacement['PID'] = wplacement['PID'].astype('int64')
        wplacement = wplacement.sort_values(by = 'Custody Start Date', ascending = False)

        #Go to current folder
        mainfolder = r'S:\DFS Data Unit\SK\Kids in Custody first and current DFS Plc SBarker'
        folders = [x for x in os.listdir(mainfolder) if re.search('\d\d$', x)]
        recentfolder = max(folders)
        mainfolder = os.path.join(mainfolder, recentfolder)

        today = dt.date.today()                     #Get todays date

        first = today.replace(day=1)
        lastMonth = first - dt.timedelta(days = 1)  #Get last month just in case

        monthfolder = ''

        for month in os.listdir(mainfolder):        #Check if any folder matches today's month
            if month.lower() == today.strftime('%B').lower() or month[:3].lower() == today.strftime('%b').lower():
                monthfolder += month
                break

        if monthfolder == '':
            for month in os.listdir(mainfolder): 
                if month.lower() == lastMonth.strftime('%B').lower() or month[:3].lower() == today.strftime('%b').lower(): #Check if last month matches if this month wasn't found   
                    monthfolder += month
                    break

        if monthfolder != '':                       #Check if this month was found in the folder
            mainfolder = os.path.join(mainfolder, monthfolder)

        else:
            p.alert(f'Search for current or last month was not found in {mainfolder}!')
            sys.exit()
            

        recentfile = max(os.listdir(mainfolder))    #Get most recent file

        filepath = os.path.join(mainfolder, recentfile) #File path to last report
        print(filepath)

        #Create DF of previous report and remove the meta data of the report at the end.
        prevdf = pd.read_excel(filepath, skipfooter = 6)

        #Add a column to see if the new report PID is in the old report and sort by New
        wplacement.insert(10, 'In Previous Report?',
                          wplacement['PID'].apply(check_prev, args=(prevdf,)), True)

        wplacement =wplacement.sort_values(by = 'In Previous Report?', ascending = False)

        #Color the New rows light orange
        wplacement1 = wplacement.style.apply(color_row, axis = 1)


        # Export to excel
        month = today.month
        day = today.day
        if len(str(day)) == 1:      #keep day format in folders as 01 rather than 1.
            day = f'0{day}'
        year = str(today.year)[-2:] #Use 20 rather than 2020
        location = fr'H:\Python_Programs\SF\Weekly\Results\KIC w Current DFS Placement FOCUS {month}{day}{year}.xlsx'
        writer = pd.ExcelWriter(location, engine = 'xlsxwriter', datetime_format = 'mm/dd/yyyy')
        wplacement1.to_excel(writer, index=False)
        writer.save()

        #Adjust column sizes
        book = load_workbook(location)
        for sheet in book.sheetnames:
            worksheet = book[sheet]
            column_size(worksheet)
        book.save(location)
        
    else:   #Otherwise use the df that was passed as an argument.
        
        ord_columns = [
            'PID', 'Name', 'Birth Date', 'Age (Years)', 'Custody Start Date','Placement Start Date',
            'Case Type','Case Owner','Supervisor','Service Area','Division','Case Number',  
            'Placement ID','Placement Type'
        ]
        wplacement = df[ord_columns].copy() #Keep the columns we need and order them.
        wplacement = wplacement[wplacement['Division']=='DFS'].copy() #Only DFS placements.
        
        wplacement['PID'] = wplacement['PID'].astype('int64')
        wplacement = wplacement.sort_values(by = 'Custody Start Date', ascending = False)

        #Go to current folder
        mainfolder = r'S:\DFS Data Unit\SK\Kids in Custody first and current DFS Plc SBarker'
        folders = [x for x in os.listdir(mainfolder) if re.search('\d\d$', x)]
        recentfolder = max(folders)
        mainfolder = os.path.join(mainfolder, recentfolder)

        today = dt.date.today()                     #Get todays date

        first = today.replace(day=1)
        lastMonth = first - dt.timedelta(days = 1)  #Get last month just in case

        monthfolder = ''

        for month in os.listdir(mainfolder):        #Check if any folder matches today's month
            if month.lower() == today.strftime('%B').lower() or month[:3].lower() == today.strftime('%b').lower():
                monthfolder += month
                break

        if monthfolder == '':
            for month in os.listdir(mainfolder): 
                if month.lower() == lastMonth.strftime('%B').lower() or month[:3].lower() == today.strftime('%b').lower(): #Check if last month matches if this month wasn't found   
                    monthfolder += month
                    break

        if monthfolder != '':                       #Check if this month was found in the folder
            mainfolder = os.path.join(mainfolder, monthfolder)

        else:
            p.alert(f'Search for current or last month was not found in {mainfolder}!')
            sys.exit()
            

        recentfile = max(os.listdir(mainfolder))    #Get most recent file

        filepath = os.path.join(mainfolder, recentfile) #File path to last report
        #print(filepath)

        #Create DF of previous report and remove the meta data of the report at the end.
        prevdf = pd.read_excel(filepath, skipfooter = 6)

        #Add a column to see if the new report PID is in the old report and sort by New
        wplacement.insert(10, 'In Previous Report?',
                          wplacement['PID'].apply(check_prev, args=(prevdf,)), True)

        wplacement =wplacement.sort_values(by = 'In Previous Report?', ascending = False)

        #Color the New rows light orange
        wplacement1 = wplacement.style.apply(color_row, axis = 1)


        # Export to excel
        month = today.month
        day = today.day
        if len(str(day)) == 1:      #keep day format in folders as 01 rather than 1.
            day = f'0{day}'
        year = str(today.year)[-2:] #Use 20 rather than 2020
        location = fr'H:\Python_Programs\SF\Weekly\Results\KIC w Current DFS Placement FOCUS {month}{day}{year}.xlsx'
        writer = pd.ExcelWriter(location, engine = 'xlsxwriter', datetime_format = 'mm/dd/yyyy')
        wplacement1.to_excel(writer, index=False)
        writer.save()

        #Adjust column sizes
        book = load_workbook(location)
        for sheet in book.sheetnames:
            worksheet = book[sheet]
            column_size(worksheet)
        book.save(location)        
        
